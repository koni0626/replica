"""
Excel(.xlsx) 読み取り・書き出し用ツール

- 依存: openpyxl
    pip install openpyxl

- 役割:
  * 指定パスの .xlsx ファイルからシートごとのテキストを抽出して返す（read_xlsx_text）
  * CSV(.csv) から .xlsx へ変換して保存する（convert_csv_to_xlsx）

注意:
- .xls (97-2003) 形式は対象外。必要なら別途変換してから使用してください。
- 数式セルは data_only=True で評価結果を優先（ブックに保存された計算結果が無い場合は空/式文字列となることがあります）。
"""
from __future__ import annotations

import os
import csv
from pathlib import Path
from typing import Optional, Tuple

try:
    import openpyxl  # type: ignore
except Exception:  # ランタイム依存を緩く
    openpyxl = None  # type: ignore


class ExcelReadError(Exception):
    """Excel 読み取り時の一般例外"""


class ExcelWriteError(Exception):
    """Excel 書き出し時の一般例外"""


def _is_xlsx(path: str) -> bool:
    return os.path.splitext(path)[1].lower() == ".xlsx"


def read_xlsx_text(
    path: str,
    limit_chars: Optional[int] = None,
    delimiter: str = "\t",
    include_sheet_names: bool = True,
) -> str:
    """
    .xlsx からテキストを抽出して返す。

    Args:
        path: 対象ファイルの絶対 or 相対パス
        limit_chars: 返却する最大文字数（None なら上限なし）。超過時は末尾を切り詰め。
        delimiter: 行整形時にセルを連結する区切り文字（既定: タブ）
        include_sheet_names: True の場合、各シートの先頭に "# Sheet: <name>" 行を付与

    Returns:
        抽出テキスト（失敗時は空文字）

    Raises:
        ExcelReadError: 依存未導入 / ファイル不正 / 読み取り失敗など致命的なときに発生
    """
    if not _is_xlsx(path):
        raise ExcelReadError(f"Unsupported extension for Excel reader: {path}")

    if openpyxl is None:
        # 依存がない環境では例外
        raise ExcelReadError(
            "openpyxl is not installed. Please `pip install openpyxl`.")

    if not os.path.exists(path):
        raise ExcelReadError(f"File not found: {path}")

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for ws in wb.worksheets:
            try:
                if include_sheet_names:
                    parts.append(f"# Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if v is None else str(v) for v in row]
                    parts.append(delimiter.join(cells))
            except Exception:
                # 個別シートで失敗しても他シート処理を継続
                continue
        text = "\n".join(parts)
        if limit_chars is not None and limit_chars > 0 and len(text) > limit_chars:
            return text[:limit_chars]
        return text
    except ExcelReadError:
        raise
    except Exception as e:
        # 解析に失敗しても原因追跡のため例外化
        raise ExcelReadError(f"Failed to read xlsx: {e}") from e


def convert_csv_to_xlsx(
    csv_path: str,
    xlsx_path: str,
    encoding: str = "cp932",
    delimiter: str = ",",
    ensure_parent: bool = True,
) -> Tuple[str, int]:
    """
    CSV ファイルを .xlsx（Excel）へ変換して保存する。

    Args:
        csv_path: 入力CSVのパス（絶対 or 相対）
        xlsx_path: 出力先 .xlsx のパス（絶対 or 相対）
        encoding: CSV のエンコーディング（既定 cp932）
        delimiter: CSV の区切り文字（既定 ,）
        ensure_parent: 出力先の親ディレクトリを自動作成するか

    Returns:
        (保存先の絶対パス, 書き込んだ行数)

    Raises:
        ExcelWriteError: 依存未導入 / 入出力エラー / 変換エラー
    """
    if openpyxl is None:
        raise ExcelWriteError("openpyxl is not installed. Please `pip install openpyxl`.")

    try:
        # パス解決
        in_path = Path(csv_path).expanduser().resolve()
        out_path = Path(xlsx_path).expanduser().resolve()
        if ensure_parent:
            out_path.parent.mkdir(parents=True, exist_ok=True)

        if not in_path.exists() or not in_path.is_file():
            raise ExcelWriteError(f"CSV not found: {in_path}")

        # CSV 読み込み → Excel へ書き出し
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        rows = 0
        with open(in_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for r in reader:
                ws.append(list(r))
                rows += 1
        wb.save(out_path)
        return str(out_path), rows
    except ExcelWriteError:
        raise
    except Exception as e:
        raise ExcelWriteError(f"Failed to convert CSV to XLSX: {e}") from e
